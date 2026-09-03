"""
eval_stage1.py — retrieval evaluation (§8).

Runs a frozen Stage 1 checkpoint (encoder -> pooling -> WordProjectionHead)
over held-out trials, ranks each query against the FULL candidate bank
(h_mid for every word OCCURRENCE across both poems, teacher_cache.py),
scores word-TYPE-level top-1/top-5 (metrics.py's label-aware accuracy —
any occurrence of the correct word counts, not just the literal matching
one), and reports against chance_level.

Also runs the --shuffle_control null distribution (controls.py's
TimeShuffledMEGTrialDataset, many seeds) — this is the actual
genuine-signal-vs-memorization check from earlier: subject-grouped CV
alone doesn't rule out the model exploiting fixed per-poem text/position
structure with the MEG signal contributing little. If real accuracy isn't
meaningfully above the shuffled null distribution, that's the result that
matters, not the raw accuracy number on its own.

Builds a qualitative prediction trace (poem, position, true word, top-1,
top-5) — this lives here, not metrics.py, because it needs the actual
poem text and word identities, which metrics.py deliberately knows
nothing about (kept pure/tensor-only so it's testable with zero real data).

NO GRU, NO GPT-2 MODEL LOADING HERE: Eval 1 is purely a Stage 1 pipeline
evaluation — encoder -> pooling -> word_head -> cosine similarity vs the
cached h_mid bank. Nothing from Stage 2 is involved.
"""

import argparse
import csv

import torch
from torch.utils.data import DataLoader

from new_dataset import MEGContinuousTrialDataset, collate_continuous_trials, _load_onsets
from splits import make_loso_splits
from pooling import pool_words
from new_controls import make_control
from metrics import cosine_similarity_matrix, topk_accuracy_label_aware, topk_predictions, chance_level, permutation_percentile
from train import POEM_TO_ID, _move_batch
from new_models import MEGEncoder, WordProjectionHead
from pooling import WordAttentionPooling


def load_stage1_checkpoint(ckpt_path: str, device):
    """Load encoder, word_head, pooling_module from a Stage 1 checkpoint."""
    ckpt = torch.load(ckpt_path, map_location=device, weights_only=False)
    pooling_mode = ckpt.get("pooling_mode", "exact")
    ckpt_subject = ckpt.get("heldout_subject", None)

    encoder = MEGEncoder().to(device)
    encoder.load_state_dict(ckpt["encoder"])
    encoder.freeze()

    word_head = WordProjectionHead(encoder.backbone_dim).to(device)
    word_head.load_state_dict(ckpt["word_head"])
    word_head.eval()

    pooling_module = WordAttentionPooling(encoder.backbone_dim).to(device)
    if "pooling" in ckpt:
        pooling_module.load_state_dict(ckpt["pooling"])
    pooling_module.eval()

    print(f"Loaded Stage 1 checkpoint: {ckpt_path}  "
          f"(epoch={ckpt.get('epoch','?')}, val_loss={ckpt.get('val_loss','?'):.4f}, "
          f"pooling_mode={pooling_mode}, heldout={ckpt_subject})")
    return encoder, word_head, pooling_module, pooling_mode, ckpt_subject


# ===========================================================================
#  Candidate bank
# ===========================================================================

def build_candidate_bank(teacher_cache: dict, poem_to_id: dict = POEM_TO_ID):
    """
    Combines h_mid across both poems into ONE candidate bank (N_total, 128),
    with a word-TYPE label per candidate (lowercased word text) for the
    label-aware top-k accuracy. teacher_cache.py itself doesn't store word
    text, so this reads the SAME onset JSON dataset.py reads
    (new_dataset.py's _load_onsets) to recover it, and asserts the word
    count matches h_mid's row count — if these ever disagree, something
    upstream (alignment, caching) is broken and this fails loudly rather
    than silently mismatching words to vectors.
    """
    vectors, word_types, poem_ids, word_pos = [], [], [], []
    for poem in ("poem1", "poem2"):
        h_mid = teacher_cache[poem]["h_mid"]
        onsets = _load_onsets(poem)
        assert len(onsets) == h_mid.shape[0], (
            f"{poem}: h_mid has {h_mid.shape[0]} rows but the onset file has {len(onsets)} words — "
            f"these must match 1:1, something upstream is inconsistent."
        )
        for pos, entry in enumerate(onsets):
            vectors.append(h_mid[pos])
            word_types.append(entry["word"].strip().lower())
            poem_ids.append(poem_to_id[poem])
            word_pos.append(pos)

    bank_vectors = torch.stack(vectors)                      # (N_total, 128)
    unique_types = sorted(set(word_types))
    type_to_id = {w: i for i, w in enumerate(unique_types)}
    bank_labels = torch.tensor([type_to_id[w] for w in word_types], dtype=torch.long)
    return bank_vectors, bank_labels, word_types, type_to_id, poem_ids, word_pos


# ===========================================================================
#  Frozen Stage 1 forward pass over a loader
# ===========================================================================

def run_stage1_forward(loader, encoder, word_head, pooling_module, pooling_mode, device):
    """
    Runs the frozen Stage 1 pipeline over every batch in loader, returns
    FLATTENED (valid words only) query vectors + their identifying info.
    A plain Python accumulation loop — dataset size is small enough that
    building simple Python lists here is clearer than trying to
    pre-allocate and index into padded tensors across variable-length batches.
    """
    encoder.eval(); word_head.eval(); pooling_module.eval()
    all_z, all_word_text, all_poem, all_word_pos, all_session = [], [], [], [], []

    with torch.no_grad():
        for batch in loader:
            batch = _move_batch(batch, device)
            z_dense = encoder(batch["meg_trial"])
            pooled, pool_valid = pool_words(
                pooling_mode, z_dense, batch["onset_samples"],
                offset_samples=batch["offset_samples"], trial_mask=batch["trial_mask"],
                attention_module=pooling_module, jitter_ms=None,
            )
            z_word = word_head(pooled)
            combined_valid = pool_valid & batch["valid_mask"]

            B, N, _ = z_word.shape
            for b in range(B):
                for i in range(N):
                    if combined_valid[b, i]:
                        all_z.append(z_word[b, i].cpu())
                        all_word_text.append(batch["word_texts"][b][i])
                        all_poem.append(batch["poem"][b])
                        all_word_pos.append(batch["word_poses"][b][i])
                        all_session.append(batch["session"][b])

    if len(all_z) == 0:
        return torch.zeros(0, encoder.backbone_dim), [], [], [], []
    return torch.stack(all_z), all_word_text, all_poem, all_word_pos, all_session


# ===========================================================================
#  Scoring
# ===========================================================================

def score_queries(z_word, word_text, bank_vectors, bank_labels, type_to_id, ks=(1, 5)):
    """
    Runs cosine similarity + label-aware top-k accuracy for a set of
    already-computed query vectors. Returns (accuracies dict, scores
    matrix, valid_mask, query_labels) — the latter three are handed back
    so callers (real run vs. shuffle control) can also build a prediction
    trace from the SAME scoring pass without recomputing it.
    """
    device = z_word.device   # query_labels MUST be created on this device explicitly —
                              # torch.tensor(...) with no device= defaults to CPU regardless
                              # of what device everything else (z_word, bank_vectors,
                              # bank_labels) is already on.
    query_labels = torch.tensor([type_to_id.get(w, -1) for w in word_text], dtype=torch.long, device=device)
    valid = query_labels >= 0   # defensive: not expected to trigger under LOSO (test words are
                                 # always from the same two poems the bank was built from), but
                                 # WOULD matter for a stimulus split evaluating held-out lines —
                                 # this guard keeps that case from silently mis-scoring instead of
                                 # requiring a stimulus-split-specific code path here.

    scores = cosine_similarity_matrix(z_word, bank_vectors)
    accs = {}
    for k in ks:
        accs[f"top{k}"] = topk_accuracy_label_aware(scores, query_labels, bank_labels, k, valid_mask=valid)
    return accs, scores, valid, query_labels


def build_prediction_trace(word_text, poem_list, word_pos_list, session_list, topk_pred_indices, bank_word_types):
    trace = []
    for i in range(len(word_text)):
        preds = [bank_word_types[idx] for idx in topk_pred_indices[i].tolist()]
        trace.append({
            "poem": poem_list[i], "word_pos": word_pos_list[i], "session": session_list[i],
            "true_word": word_text[i], "top1_pred": preds[0], "topk_preds": preds,
            "correct_top1": preds[0] == word_text[i], "correct_topk": word_text[i] in preds,
        })
    return trace


def save_prediction_trace_excel(trace, path):
    """
    Color-coded Excel workbook: one sheet per poem, rows sorted (session,
    word_pos) so each sheet reads like the poem being presented, repeated
    across sessions. Green = correct top-1, light green = correct within
    top-5 but not top-1, red = miss (true word isn't even in the top-5).

    Needs openpyxl (pip install openpyxl) — not one of this project's
    existing dependencies until now.
    """
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font
    from openpyxl.utils import get_column_letter

    GREEN = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    LIGHT_GREEN = PatternFill(start_color="E2F0D9", end_color="E2F0D9", fill_type="solid")
    RED = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    wb = Workbook()
    wb.remove(wb.active)   # one sheet per poem instead of the default blank sheet

    headers = ["session", "word_pos", "true_word", "top1_pred", "top5_preds", "hit_top1", "hit_top5"]
    for poem in sorted(set(row["poem"] for row in trace)):
        ws = wb.create_sheet(title=poem)
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True)

        rows = sorted((r for r in trace if r["poem"] == poem), key=lambda r: (r["session"], r["word_pos"]))
        for r in rows:
            ws.append([
                r["session"], r["word_pos"], r["true_word"], r["top1_pred"],
                ",".join(r["topk_preds"]), r["correct_top1"], r["correct_topk"],
            ])
            fill = GREEN if r["correct_top1"] else (LIGHT_GREEN if r["correct_topk"] else RED)
            for col in range(1, len(headers) + 1):
                ws.cell(row=ws.max_row, column=col).fill = fill

        for i, width in enumerate([9, 9, 14, 14, 45, 9, 9], start=1):
            ws.column_dimensions[get_column_letter(i)].width = width

    wb.save(path)
    n_poems = len(wb.sheetnames)
    print(f"Saved color-coded prediction trace -> {path}  ({n_poems} sheet(s): {', '.join(wb.sheetnames)})")


def print_prediction_trace(trace, n=20):
    print(f"\n{'poem':7s} {'pos':4s} {'true':14s} {'top1':14s} {'top5':50s} {'hit1':5s} {'hit5':5s}")
    for row in trace[:n]:
        topk_str = ",".join(row["topk_preds"])
        print(f"{row['poem']:7s} {row['word_pos']:<4d} {row['true_word']:14s} {row['top1_pred']:14s} "
              f"{topk_str:50s} {str(row['correct_top1']):5s} {str(row['correct_topk']):5s}")
    if len(trace) > n:
        print(f"... ({len(trace) - n} more rows, use --save_trace_csv to see all)")


def save_prediction_trace_csv(trace, path):
    fieldnames = ["poem", "session", "word_pos", "true_word", "top1_pred", "topk_preds", "correct_top1", "correct_topk"]
    with open(path, "w", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        for row in trace:
            r = dict(row)
            r["topk_preds"] = "|".join(r["topk_preds"])
            writer.writerow(r)
    print(f"Saved prediction trace ({len(trace)} rows) -> {path}")


# ===========================================================================
#  Main
# ===========================================================================

def main(args):
    device = torch.device("cuda" if torch.cuda.is_available() else "cpu")

    encoder, word_head, pooling_module, pooling_mode, ckpt_subject = load_stage1_checkpoint(
        args.stage1_checkpoint_path, device
    )
    if ckpt_subject is not None and ckpt_subject != args.heldout_subject:
        print(f"  WARNING: checkpoint was trained with heldout_subject={ckpt_subject!r}, "
              f"but --heldout_subject={args.heldout_subject!r} was passed. Proceeding, but double-check this.")

    teacher_cache = torch.load(args.teacher_cache_path, weights_only=False)
    bank_vectors, bank_labels, bank_word_types, type_to_id, _, _ = build_candidate_bank(teacher_cache)
    bank_vectors, bank_labels = bank_vectors.to(device), bank_labels.to(device)
    print(f"Candidate bank: {bank_vectors.shape[0]} occurrences, {len(type_to_id)} unique word types")

    splits = make_loso_splits(args.heldout_subject)
    test_ds = MEGContinuousTrialDataset(splits["test"]["trials"], word_filter=splits["test"]["word_filter"],
                                         meg_base=args.meg_base)
    test_loader = DataLoader(test_ds, batch_size=args.batch_size, shuffle=False, collate_fn=collate_continuous_trials)

    z_word, word_text, poem_list, word_pos_list, session_list = run_stage1_forward(
        test_loader, encoder, word_head, pooling_module, pooling_mode, device
    )
    z_word = z_word.to(device)
    accs, scores, valid, query_labels = score_queries(z_word, word_text, bank_vectors, bank_labels, type_to_id, ks=(1, 5))

    n_candidates = bank_vectors.shape[0]
    n_queries = valid.sum().item()
    print(f"\n=== Eval 1 (retrieval) — heldout={args.heldout_subject} ===")
    print(f"n_queries={n_queries}  n_candidates={n_candidates}")
    for k in (1, 5):
        print(f"top{k}: {accs[f'top{k}']*100:.2f}%  (chance: {chance_level(n_candidates, k)*100:.2f}%)")

    preds5 = topk_predictions(scores, k=5)
    trace = build_prediction_trace(word_text, poem_list, word_pos_list, session_list, preds5, bank_word_types)
    print_prediction_trace(trace, n=args.trace_rows)
    if args.save_trace_csv:
        save_prediction_trace_csv(trace, args.save_trace_csv)
    if args.save_trace_excel:
        save_prediction_trace_excel(trace, args.save_trace_excel)

    if args.shuffle_control:
        print(f"\n=== shuffle control ({args.num_shuffle_seeds} seeds) ===")
        null_top1, null_top5 = [], []
        for seed in range(args.num_shuffle_seeds):
            shuf_ds = make_control(test_ds, "shuffle_time", seed=seed)
            shuf_loader = DataLoader(shuf_ds, batch_size=args.batch_size, shuffle=False,
                                      collate_fn=collate_continuous_trials)
            z_s, wt_s, _, _, _ = run_stage1_forward(shuf_loader, encoder, word_head, pooling_module, pooling_mode, device)
            if z_s.shape[0] == 0:
                continue
            accs_s, _, _, _ = score_queries(z_s.to(device), wt_s, bank_vectors, bank_labels, type_to_id, ks=(1, 5))
            null_top1.append(accs_s["top1"])
            null_top5.append(accs_s["top5"])

        p1 = permutation_percentile(accs["top1"], null_top1)
        p5 = permutation_percentile(accs["top5"], null_top5)
        print(f"null top1: mean={sum(null_top1)/len(null_top1)*100:.2f}%  "
              f"real={accs['top1']*100:.2f}%  fraction of null >= real={p1:.3f}")
        print(f"null top5: mean={sum(null_top5)/len(null_top5)*100:.2f}%  "
              f"real={accs['top5']*100:.2f}%  fraction of null >= real={p5:.3f}")
        print("(smaller 'fraction of null >= real' means the real result sits further into the "
              "tail of the shuffled-MEG null distribution — more consistent with genuine signal.)")

    if args.zero_control:
        print(f"\n=== zero-MEG control ===")
        zero_ds = make_control(test_ds, "zero")
        zero_loader = DataLoader(zero_ds, batch_size=args.batch_size, shuffle=False, collate_fn=collate_continuous_trials)
        z_z, wt_z, _, _, _ = run_stage1_forward(zero_loader, encoder, word_head, pooling_module, pooling_mode, device)
        if z_z.shape[0] == 0:
            print("  (no valid queries under zero-MEG — nothing to score)")
        else:
            accs_z, _, _, _ = score_queries(z_z.to(device), wt_z, bank_vectors, bank_labels, type_to_id, ks=(1, 5))
            print(f"  top1={accs_z['top1']*100:.2f}%  top5={accs_z['top5']*100:.2f}%  "
                  f"(real MEG: top1={accs['top1']*100:.2f}%, top5={accs['top5']*100:.2f}%)")
            print("  (zero-MEG replaces the signal entirely with zeros, so every word's query collapses "
                  "toward the same fixed vector — accuracy near chance here is expected and reassuring; "
                  "if it's NOT near chance, something is leaking non-MEG information, e.g. through "
                  "pooling window position alone.)")


# ===========================================================================
#  --dry_run: fully synthetic, no real files needed at all.
# ===========================================================================

def run_dry_run(device):
    print("=== eval_stage1.py DRY RUN (fully synthetic) ===\n")
    from new_models import MEGEncoder, WordProjectionHead
    from pooling import WordAttentionPooling

    encoder = MEGEncoder().to(device)
    word_head = WordProjectionHead(encoder.backbone_dim).to(device)
    pooling_module = WordAttentionPooling(encoder.backbone_dim).to(device)

    # Fake candidate bank: 6 occurrences, 3 unique word types (2 each) —
    # small enough to hand-verify, big enough to exercise label-aware matching.
    bank_vectors = torch.randn(6, encoder.backbone_dim).to(device)
    bank_word_types = ["cat", "dog", "cat", "bird", "dog", "bird"]
    type_to_id = {"cat": 0, "dog": 1, "bird": 2}
    bank_labels = torch.tensor([type_to_id[w] for w in bank_word_types]).to(device)

    # Fake dataset via __new__ + manual _items — same pattern controls.py's own dry run uses.
    fake_ds = MEGContinuousTrialDataset.__new__(MEGContinuousTrialDataset)
    fake_ds._items = [{
        "meg_trial": torch.randn(155, 60) * 0.1,
        "onset_samples": torch.tensor([5, 15, 25]),
        "offset_samples": torch.tensor([10, 20, 30]),
        "valid_mask": torch.tensor([True, True, True]),
        "word_texts": ["cat", "dog", "bird"], "word_poses": [0, 1, 2],
        "poem": "poem1", "subject": "sub-fake", "session": 0,
    }, {
        "meg_trial": torch.randn(155, 60) * 0.1,
        "onset_samples": torch.tensor([5, 15, 25]),
        "offset_samples": torch.tensor([10, 20, 30]),
        "valid_mask": torch.tensor([True, True, True]),
        "word_texts": ["dog", "cat", "bird"], "word_poses": [0, 1, 2],
        "poem": "poem1", "subject": "sub-fake2", "session": 0,
    }]

    loader = DataLoader(fake_ds, batch_size=2, shuffle=False, collate_fn=collate_continuous_trials)
    z_word, word_text, poem_list, word_pos_list, session_list = run_stage1_forward(
        loader, encoder, word_head, pooling_module, "wide", device
    )
    assert z_word.shape == (6, encoder.backbone_dim), f"expected 6 valid queries, got {z_word.shape}"
    assert session_list == [0, 0, 0, 0, 0, 0], f"both fake trials use session=0, got {session_list}"
    print(f"[OK] run_stage1_forward produced {z_word.shape[0]} query vectors (with session info) from 2 fake trials")

    accs, scores, valid, query_labels = score_queries(z_word.to(device), word_text, bank_vectors, bank_labels, type_to_id)
    assert scores.shape == (6, 6)
    assert valid.all(), "all 6 fake words are real bank vocabulary — none should be marked invalid"
    print(f"[OK] score_queries: top1={accs['top1']*100:.1f}%  top5={accs['top5']*100:.1f}%")

    preds = topk_predictions(scores, k=3)
    trace = build_prediction_trace(word_text, poem_list, word_pos_list, session_list, preds, bank_word_types)
    assert len(trace) == 6 and all(len(row["topk_preds"]) == 3 for row in trace)
    assert all("session" in row for row in trace)
    print(f"[OK] build_prediction_trace produced {len(trace)} rows, each with session info")

    # Excel export itself — write to a temp file, confirm it's readable back
    # and has the color fill actually applied (not just written silently wrong).
    import tempfile, os as _os
    from openpyxl import load_workbook
    with tempfile.TemporaryDirectory() as tmpdir:
        xlsx_path = _os.path.join(tmpdir, "trace_test.xlsx")
        save_prediction_trace_excel(trace, xlsx_path)
        wb = load_workbook(xlsx_path)
        assert "poem1" in wb.sheetnames
        ws = wb["poem1"]
        assert ws.max_row == 1 + len(trace), f"expected header + {len(trace)} data rows, got {ws.max_row}"
        first_data_row_fill = ws.cell(row=2, column=1).fill.start_color.rgb
        assert first_data_row_fill in ("00C6EFCE", "00E2F0D9", "00FFC7CE"), \
            f"expected one of the three defined fill colors, got {first_data_row_fill}"
    print("[OK] save_prediction_trace_excel writes a readable .xlsx with color fills correctly applied")

    # Shuffle control end-to-end on the fake dataset (real controls.py logic).
    shuf_ds = make_control(fake_ds, "shuffle_time", seed=1)
    shuf_loader = DataLoader(shuf_ds, batch_size=2, shuffle=False, collate_fn=collate_continuous_trials)
    z_s, wt_s, _, _, _ = run_stage1_forward(shuf_loader, encoder, word_head, pooling_module, "wide", device)
    assert z_s.shape[0] == 6, "shuffle control must not change word count, only onset assignment"
    accs_s, _, _, _ = score_queries(z_s.to(device), wt_s, bank_vectors, bank_labels, type_to_id)
    p = permutation_percentile(accs["top1"], [accs_s["top1"], accs["top1"] * 0.5])
    assert 0.0 <= p <= 1.0
    print(f"[OK] shuffle control runs end-to-end (shuffled top1={accs_s['top1']*100:.1f}%), "
          f"permutation_percentile in valid range")

    # Zero-MEG control: every word's query should collapse toward the SAME
    # fixed vector, since the input signal is identical (all zeros) for
    # every word regardless of which word it actually is.
    zero_ds = make_control(fake_ds, "zero")
    zero_loader = DataLoader(zero_ds, batch_size=2, shuffle=False, collate_fn=collate_continuous_trials)
    z_zero, wt_zero, _, _, _ = run_stage1_forward(zero_loader, encoder, word_head, pooling_module, "wide", device)
    assert z_zero.shape[0] == 6, "zero control must not change word count"
    pairwise_diff = (z_zero - z_zero[0]).abs().max().item()
    assert pairwise_diff < 1e-4, (
        f"zero-MEG queries should all collapse to the same vector (deterministic constant input), "
        f"but max pairwise difference was {pairwise_diff}"
    )
    accs_zero, _, _, _ = score_queries(z_zero.to(device), wt_zero, bank_vectors, bank_labels, type_to_id)
    print(f"[OK] zero-MEG control: all {z_zero.shape[0]} queries collapsed to the same vector "
          f"(max pairwise diff={pairwise_diff:.2e}), top1={accs_zero['top1']*100:.1f}%")

    print("\n=== DRY RUN PASSED ===")


def build_arg_parser():
    p = argparse.ArgumentParser(description="Eval 1: retrieval evaluation (§8).")
    p.add_argument("--stage1_checkpoint_path", type=str, default=None)
    p.add_argument("--heldout_subject", type=str, default="sub-01")
    p.add_argument("--batch_size", type=int, default=4)
    p.add_argument("--teacher_cache_path", type=str, default="teacher_cache.pt")
    p.add_argument("--meg_base", type=str, default=None)
    p.add_argument("--trace_rows", type=int, default=20, help="How many prediction-trace rows to print to console.")
    p.add_argument("--save_trace_csv", type=str, default=None, help="If set, save the FULL trace to this CSV path.")
    p.add_argument("--save_trace_excel", type=str, default=None,
                    help="If set, save a color-coded .xlsx (needs `pip install openpyxl`) — "
                         "green=correct top-1, light green=correct top-5, red=miss, one sheet per poem.")
    p.add_argument("--shuffle_control", action="store_true")
    p.add_argument("--num_shuffle_seeds", type=int, default=20)
    p.add_argument("--zero_control", action="store_true")
    p.add_argument("--dry_run", action="store_true")
    return p


if __name__ == "__main__":
    args = build_arg_parser().parse_args()
    _device = torch.device("cuda" if torch.cuda.is_available() else "cpu")
    if args.dry_run:
        run_dry_run(_device)
    else:
        if args.stage1_checkpoint_path is None:
            raise ValueError("--stage1_checkpoint_path is required outside --dry_run")
        main(args)
